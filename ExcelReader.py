from openpyxl import load_workbook
import re


def go():
    wb = load_workbook(filename="excel_data.xlsx", data_only=True)
    print()
    ws = wb["Arkusz1"]
    decisions = [["B5:J11", 0, 1],
              ["B16:G22", 0, 2],
              ["B29:N37", 1, 3],
              ["B42:K47", 2, 4],
              ["B52:N57", 2, 3],
              ["B62:M70", 1, 4],
              ["B75:R87", 3, 5],
              ["B91:O103", 3, 6],
              ["B108:O119", 4, 6]]
    cirerias = ["B137:D143",
                "B146:D154",
                "G137:I142",
                "G146:I157",
                "B160:D169",
                "G160:I176",
                "L160:N173"]
    coefficients = "B191:D197"
    feedback_decisions = "B127:K132"
    feedback_headers = "A127:A132"

    coefficients_dict = {i: {j: ws[coefficients][i][j].value for j in range(len(ws[coefficients][i]))}
                         for i in range(len(ws[coefficients]))}

    decsions_dict = {}
    for dec in decisions:
        matrix = ws[dec[0]]
        if dec[1] not in decsions_dict:
            decsions_dict[dec[1]] = {}
        if dec[2] not in decsions_dict[dec[1]]:
            decsions_dict[dec[1]][dec[2]] = {}
        for i in range(len(matrix)):
            for j in range(len(matrix[0])):
                if i+1 not in decsions_dict[dec[1]][dec[2]]:
                    decsions_dict[dec[1]][dec[2]][i+1] = {}
                decsions_dict[dec[1]][dec[2]][i+1][j+1] = matrix[i][j].value

    criterias_dict = {}
    for i, val in enumerate(cirerias):
        if i not in criterias_dict:
            criterias_dict[i] = {}
        for j, k in enumerate(ws[val]):
            criterias_dict[i][j+1] = [x.value for x in k]

    feedbacks_dict = {}
    for i, val in enumerate(ws[feedback_headers]):
        read = val[0].value
        route = re.findall(r'\d+', read)
        route = [int(x) for x in route]
        if route[0] not in feedbacks_dict:
            feedbacks_dict[route[0]] = {}
        if route[1] not in feedbacks_dict[route[0]]:
            feedbacks_dict[route[0]][route[1]] = []
        for j, elem in enumerate(ws[feedback_decisions][i]):
            read = elem.value
            if read is not None:
                feedbacks_dict[route[0]][route[1]].append(int(re.findall(r'\d+', read)[1]))

    return decsions_dict, criterias_dict, feedbacks_dict, coefficients_dict
